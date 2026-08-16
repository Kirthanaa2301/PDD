import os
import openpyxl
import yaml
import hashlib

def validate_excel_report(file_path, expected_prefix, expected_count=300):
    print(f"Validating: {os.path.basename(file_path)}...")
    if not os.path.exists(file_path):
        print(f"  [ERROR] File does not exist: {file_path}")
        return False
        
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"  [ERROR] Failed to load workbook: {e}")
        return False
        
    # Check sheet existence
    if "Executive Summary" not in wb.sheetnames:
        print("  [ERROR] Missing sheet 'Executive Summary'")
        return False
    if "Test Cases Detail" not in wb.sheetnames:
        print("  [ERROR] Missing sheet 'Test Cases Detail'")
        return False
        
    ws_det = wb["Test Cases Detail"]
    
    # Check headers
    headers = [cell.value for cell in ws_det[1]]
    required_cols = ["Test ID", "Test Case Name", "Module", "Category", "Endpoint/Screen", "Method/Action", "Priority", "Status", "Duration (ms)", "Actual Result"]
    for col in required_cols:
        if col not in headers:
            print(f"  [ERROR] Required column '{col}' is missing from header row")
            return False
            
    # Check rows and values
    ids = []
    descriptions = []
    status_idx = headers.index("Status") + 1
    
    row_count = 0
    for row in range(2, ws_det.max_row + 1):
        tc_id = ws_det.cell(row=row, column=1).value
        name = ws_det.cell(row=row, column=2).value
        status_val = ws_det.cell(row=row, column=status_idx).value
        
        if tc_id is None:
            # Reached empty rows
            continue
            
        row_count += 1
        ids.append(tc_id)
        descriptions.append(name)
        
        # Verify status is exactly PASS
        if status_val != "PASS":
            print(f"  [ERROR] Row {row} ID {tc_id} has status '{status_val}', expected 'PASS'")
            return False
            
        # Verify Test ID format and prefix
        expected_id = f"{expected_prefix}-{str(row_count).zfill(3)}"
        if tc_id != expected_id:
            print(f"  [ERROR] Row {row} ID is '{tc_id}', expected sequential '{expected_id}'")
            return False

    if row_count != expected_count:
        print(f"  [ERROR] Detailed sheet contains {row_count} test cases, expected exactly {expected_count}")
        return False
        
    # Check duplicate IDs
    if len(ids) != len(set(ids)):
        print("  [ERROR] Duplicate Test IDs found in sheet")
        return False
        
    # Check duplicate descriptions
    if len(descriptions) != len(set(descriptions)):
        print("  [ERROR] Duplicate Test Case Names (descriptions) found in sheet")
        seen = set()
        dupes = []
        for name in descriptions:
            if name in seen:
                dupes.append(name)
            seen.add(name)
        print(f"  First few duplicate descriptions: {dupes[:3]}")
        return False
        
    print(f"  [SUCCESS] Workbook verified: exactly {row_count} sequential, unique PASS test cases.")
    return True

def validate_executive_report(file_path):
    print(f"Validating Executive Report: {os.path.basename(file_path)}...")
    if not os.path.exists(file_path):
        print(f"  [ERROR] File does not exist: {file_path}")
        return False
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"  [ERROR] Failed to load workbook: {e}")
        return False
        
    ws = wb["Executive Summary"]
    
    # Check cells for total stats
    found_total = False
    for r in range(4, 20):
        val_a = ws.cell(row=r, column=1).value
        if val_a == "TOTAL":
            tot = ws.cell(row=r, column=2).value
            passed = ws.cell(row=r, column=3).value
            failed = ws.cell(row=r, column=4).value
            blocked = ws.cell(row=r, column=5).value
            not_run = ws.cell(row=r, column=6).value
            
            if tot == 1200 and passed == 1200 and failed == 0 and blocked == 0 and not_run == 0:
                found_total = True
            else:
                print(f"  [ERROR] Executive Report summary table TOTAL values check failed: Total={tot}, Passed={passed}, Failed={failed}, Blocked={blocked}, Not Run={not_run}")
            break
            
    if not found_total:
        print("  [ERROR] Executive Report summary table total row check failed. Expected Total=1200, Passed=1200, Failed=0, Blocked=0, Not Run=0.")
        return False
        
    print("  [SUCCESS] Executive report validated.")
    return True

def validate_github_workflow():
    print("Validating GitHub Actions workflow...")
    workflow_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".github", "workflows", "qa.yml")
    if not os.path.exists(workflow_path):
        print("  [WARNING] GitHub Actions workflow does not exist yet.")
        return False
        
    try:
        with open(workflow_path, "r") as f:
            yaml.safe_load(f)
        print("  [SUCCESS] GitHub Actions workflow has valid YAML syntax.")
        return True
    except Exception as e:
        print(f"  [ERROR] Invalid YAML in GitHub Actions workflow: {e}")
        return False

def validate_global_uniqueness(reports_dir):
    print("Validating global uniqueness across all 4 test suites...")
    all_rows = []
    ids = []
    names = []
    signatures = []
    
    col_data = {
        "Test Case Name": [],
        "Module": [],
        "Category": [],
        "Endpoint/Screen": [],
        "Method/Action": [],
        "Actual Result": [],
        "Preconditions": [],
        "Test Steps": [],
        "Expected Result": []
    }
    
    files = [
        ("Selenium_300_Test_Report.xlsx", "TC-SEL"),
        ("Appium_300_Test_Report.xlsx", "TC-APP"),
        ("Security_300_Test_Report.xlsx", "TC-SEC"),
        ("Load_300_Test_Report.xlsx", "TC-LOAD")
    ]
    
    for filename, prefix in files:
        filepath = os.path.join(reports_dir, filename)
        if not os.path.exists(filepath):
            print(f"  [ERROR] File does not exist for uniqueness check: {filename}")
            return False
        
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb["Test Cases Detail"]
        except Exception as e:
            print(f"  [ERROR] Failed to load {filename}: {e}")
            return False
            
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        required = ["Test ID", "Test Case Name", "Module", "Category", "Endpoint/Screen", "Method/Action", "Priority", "Status", "Duration (ms)", "Actual Result", "Preconditions", "Test Steps", "Expected Result"]
        for col in required:
            if col not in headers:
                print(f"  [ERROR] Required column {col} missing in {filename}")
                return False
                
        id_idx = headers.index("Test ID") + 1
        name_idx = headers.index("Test Case Name") + 1
        mod_idx = headers.index("Module") + 1
        cat_idx = headers.index("Category") + 1
        end_idx = headers.index("Endpoint/Screen") + 1
        met_idx = headers.index("Method/Action") + 1
        prio_idx = headers.index("Priority") + 1
        act_idx = headers.index("Actual Result") + 1
        prec_idx = headers.index("Preconditions") + 1
        steps_idx = headers.index("Test Steps") + 1
        exp_idx = headers.index("Expected Result") + 1
        
        for r in range(2, ws.max_row + 1):
            tc_id = ws.cell(row=r, column=id_idx).value
            if tc_id is None:
                continue
                
            tc_name = ws.cell(row=r, column=name_idx).value
            mod = ws.cell(row=r, column=mod_idx).value
            cat = ws.cell(row=r, column=cat_idx).value
            end = ws.cell(row=r, column=end_idx).value
            met = ws.cell(row=r, column=met_idx).value
            prio = ws.cell(row=r, column=prio_idx).value
            act = ws.cell(row=r, column=act_idx).value
            prec = ws.cell(row=r, column=prec_idx).value
            steps_val = ws.cell(row=r, column=steps_idx).value
            exp = ws.cell(row=r, column=exp_idx).value
            
            row_vals = tuple(str(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1))
            all_rows.append(row_vals)
            
            ids.append(str(tc_id).strip())
            names.append(str(tc_name).strip().lower())
            
            col_data["Test Case Name"].append(str(tc_name).strip())
            col_data["Module"].append(str(mod).strip())
            col_data["Category"].append(str(cat).strip())
            col_data["Endpoint/Screen"].append(str(end).strip())
            col_data["Method/Action"].append(str(met).strip())
            col_data["Actual Result"].append(str(act).strip())
            col_data["Preconditions"].append(str(prec).strip())
            col_data["Test Steps"].append(str(steps_val).strip())
            col_data["Expected Result"].append(str(exp).strip())
            
            sig_str = f"{str(tc_name).strip()}|{str(mod).strip()}|{str(cat).strip()}|{str(end).strip()}|{str(met).strip()}|{str(prio).strip()}|{str(act).strip()}"
            sig_hash = hashlib.sha256(sig_str.encode('utf-8')).hexdigest()
            signatures.append(sig_hash)
            
    if len(all_rows) != 1200:
        print(f"  [ERROR] Total combined test cases row count is {len(all_rows)}, expected exactly 1200.")
        return False
        
    if len(ids) != len(set(ids)):
        print("  [ERROR] Duplicate Test IDs found across the 4 suites!")
        seen = set()
        dupes = []
        for x in ids:
            if x in seen: dupes.append(x)
            seen.add(x)
        print(f"  Duplicated IDs: {dupes[:5]}")
        return False
        
    if len(names) != len(set(names)):
        print("  [ERROR] Duplicate Test Case Names (descriptions) found across the 4 suites!")
        seen = set()
        dupes = []
        for x in names:
            if x in seen: dupes.append(x)
            seen.add(x)
        print(f"  Duplicated names: {dupes[:5]}")
        return False
        
    if len(signatures) != len(set(signatures)):
        print("  [ERROR] Duplicate scenario signatures found across the 4 suites!")
        seen = set()
        dupes = []
        for i, sig in enumerate(signatures):
            if sig in seen:
                dupes.append(all_rows[i][0])
            seen.add(sig)
        print(f"  Duplicated signatures in test IDs: {dupes[:5]}")
        return False
        
    if len(all_rows) != len(set(all_rows)):
        print("  [ERROR] Duplicate complete rows found across the 4 suites!")
        seen = set()
        dupes = []
        for row in all_rows:
            if row in seen: dupes.append(row[0])
            seen.add(row)
        print(f"  Duplicated complete rows in test IDs: {dupes[:5]}")
        return False

    # Calculate and display duplication/repetition percentages
    print("\n" + "="*58)
    print("        QA DATA QUALITY CONTENT REPETITION REPORT")
    print("="*58)
    print(f"{'Column Name':<28} | {'Unique Count':<12} | {'Repetition %':<12}")
    print("-"*58)
    
    excessive_repetition = False
    for col_name, values in col_data.items():
        total = len(values)
        unique = len(set(values))
        rep_pct = (1 - (unique / total)) * 100
        print(f"{col_name:<28} | {unique:<12} | {rep_pct:.1f}%")
        
        # Descriptive columns should not have high repetition
        descriptive_cols = ["Test Case Name", "Method/Action", "Actual Result", "Preconditions", "Test Steps", "Expected Result"]
        if col_name in descriptive_cols and rep_pct > 5.0:
            print(f"  [WARNING] Column '{col_name}' has excessive repetition ({rep_pct:.1f}%)!")
            excessive_repetition = True
            
    print("="*58)
    
    if excessive_repetition:
        print("  [ERROR] Descriptive columns show duplicated templates. Content verification failed.")
        return False
        
    print("  [SUCCESS] Global uniqueness validated: 1,200 unique IDs, names, complete rows, and scenario signatures.")
    return True

def run_all_validations():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    reports_dir = os.path.join(os.path.dirname(base_dir), "reports")
    
    sel_path = os.path.join(reports_dir, "Selenium_300_Test_Report.xlsx")
    app_path = os.path.join(reports_dir, "Appium_300_Test_Report.xlsx")
    sec_path = os.path.join(reports_dir, "Security_300_Test_Report.xlsx")
    load_path = os.path.join(reports_dir, "Load_300_Test_Report.xlsx")
    exec_path = os.path.join(reports_dir, "QA_Executive_Report.xlsx")
    exec_1200_path = os.path.join(reports_dir, "QA_1200_Test_Executive_Report.xlsx")
    
    success = True
    success &= validate_excel_report(sel_path, "TC-SEL", 300)
    success &= validate_excel_report(app_path, "TC-APP", 300)
    success &= validate_excel_report(sec_path, "TC-SEC", 300)
    success &= validate_excel_report(load_path, "TC-LOAD", 300)
    success &= validate_executive_report(exec_path)
    success &= validate_executive_report(exec_1200_path)
    success &= validate_github_workflow()
    success &= validate_global_uniqueness(reports_dir)
    
    if success:
        print("\n[CONGRATULATIONS] All 1,200 QA validation checks passed successfully!")
        return 0
    else:
        print("\n[FAILURE] QA validation checks failed. See errors above.")
        return 1

if __name__ == "__main__":
    import sys
    sys.exit(run_all_validations())
